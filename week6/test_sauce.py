#odev2

from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium import webdriver
from pathlib import Path
from time import sleep
import xlsxwriter
import constants
import openpyxl
import pytest
import pandas

class Test_SauceDemo:

    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(constants.URL)
        self.folderPath = "week6/screenshots"
        Path(self.folderPath).mkdir(exist_ok=True)
    
    def teardown_method(self): 
        self.driver.quit()

    def waitForElementVisible(self, locator, timeout=3):
        WebDriverWait(self.driver, timeout).until(ec.visibility_of_element_located(locator))
    
    def successfulLogin(self):
        self.waitForElementVisible((By.ID, "user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        usernameInput.send_keys("standard_user")
        self.waitForElementVisible((By.ID, "password"))
        passwordInput = self.driver.find_element(By.ID, "password")
        passwordInput.send_keys("secret_sauce")
        self.waitForElementVisible((By.ID, "login-button"))
        self.driver.find_element(By.ID, "login-button").click()

    def getUsernameAndPassword(fileName, sheetName):
        excelFile = openpyxl.load_workbook(f"week6/data/{fileName}.xlsx")
        totalRows = excelFile[sheetName].max_row
        data = []
        for i in range(2,totalRows+1):
            username = excelFile[sheetName].cell(i,1).value
            password = excelFile[sheetName].cell(i,2).value
            tupleData = (username, password)
            data.append(tupleData)

        return data
    
    def getProductName(fileName, sheetName):
        excelFile = openpyxl.load_workbook(f"week6/data/{fileName}.xlsx")
        totalRows = excelFile[sheetName].max_row
        data = []
        for i in range(2,totalRows+1):
            productName = str(excelFile[sheetName].cell(i,1).value).lower().replace(" ", "-")
            data.append(productName)

        return data

    @pytest.mark.parametrize("username,password", getUsernameAndPassword("valid_login", "Sheet1"))
    def test_valid_login(self, username, password):
        self.waitForElementVisible((By.ID, "user-name")) 
        usernameInput = self.driver.find_element(By.ID, "user-name")
        usernameInput.send_keys(username)
        self.waitForElementVisible((By.ID, "password"))    
        passwordInput = self.driver.find_element(By.ID, "password")         
        passwordInput.send_keys(password)
        self.waitForElementVisible((By.ID, "login-button"))
        self.driver.find_element(By.ID, "login-button").click()
        sleep(0.2) #200 ms delay for loading image
        self.driver.save_screenshot(f"{self.folderPath}/test-valid-login-{username}-{password}.png")
        assert self.driver.current_url == constants.MAIN_PAGE

    @pytest.mark.parametrize("username,password", getUsernameAndPassword("invalid_login", "Sheet1"))
    def test_invalid_login(self, username, password):
        self.waitForElementVisible((By.ID, "user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        usernameInput.send_keys(username)
        self.waitForElementVisible((By.ID, "password"))
        passwordInput = self.driver.find_element(By.ID, "password")
        passwordInput.send_keys(password)
        self.waitForElementVisible((By.ID, "login-button"))
        self.driver.find_element(By.ID, "login-button").click()
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        errorButton = self.driver.find_element(By.CLASS_NAME, "error-button")
        assert errorButton.is_displayed

    @pytest.mark.parametrize("productName", getProductName("shop_list", "Sheet1"))
    def test_addToCart_1by1(self, productName):
        self.successfulLogin()
        self.driver.find_element(By.XPATH, f"//*[@id='add-to-cart-{productName}']").click()
        self.driver.find_element(By.XPATH, "//*[@id='shopping_cart_container']/a").click()
        self.waitForElementVisible((By.ID, "checkout"))
        self.driver.find_element(By.XPATH, "//*[@id='checkout']").click()
        self.waitForElementVisible((By.ID, "first-name"))
        self.driver.find_element(By.XPATH, "//*[@id='first-name']").send_keys("Hakan Emin")
        self.waitForElementVisible((By.ID, "last-name"))
        self.driver.find_element(By.XPATH, "//*[@id='last-name']").send_keys("Ünal")
        self.waitForElementVisible((By.ID, "postal-code"))
        self.driver.find_element(By.XPATH, "//*[@id='postal-code']").send_keys("06")
        self.waitForElementVisible((By.ID, "continue"))
        self.driver.find_element(By.XPATH, "//*[@id='continue']").click()
        self.driver.save_screenshot(f"{self.folderPath}/test-addToCart-{productName}.png")
        self.waitForElementVisible((By.ID, "finish"))
        self.driver.find_element(By.XPATH, "//*[@id='finish']").click()
        assert self.driver.find_element(By.XPATH, "//*[@id='checkout_complete_container']/h2").text == "Thank you for your order!"
        
    def test_setProductNameAndPriceToExcel(self):
        self.successfulLogin()
        nameElements = self.driver.find_elements(By.CLASS_NAME, "inventory_item_name")
        priceElements = self.driver.find_elements(By.CLASS_NAME, "inventory_item_price")

        productNames = []
        for i in range(len(nameElements)):
            productNames.append(nameElements[i].text)

        productPrices = []
        for i in range(len(priceElements)):
            productPrices.append(priceElements[i].text)

        dataFrame = pandas.DataFrame({'product_name': productNames, 'product_price': productPrices})
        excelWriter = pandas.ExcelWriter('./week6/data/sauce_products_and_their_prices.xlsx', engine='xlsxwriter')
        dataFrame.to_excel(excelWriter, sheet_name='Sheet1', index=False)
        excelWriter.close()

        self.driver.find_element(By.XPATH, "//*[@id='react-burger-menu-btn']").click()
        self.waitForElementVisible((By.ID, "logout_sidebar_link"))
        self.driver.find_element(By.XPATH, "//*[@id='logout_sidebar_link']").click()
        self.driver.save_screenshot(f"{self.folderPath}/test-logout.png")
        assert self.driver.current_url == constants.URL
        
 

        

        
        

        



